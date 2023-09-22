import datetime
import re
import csv
import ast
import openpyxl

def contiene_digitos(nombre):
    for caracter in nombre:
        if caracter.isdigit():
            return True
    return False

notas = dict()

try:
    with open("notas.csv","r") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        for row in lector:
            folio = row[0]
            fecha_str = row[1]
            cliente = row[2]
            rfc = row[3]
            correo = row[4]
            detalle_str = row[5]
            monto = row[6]
            estado =row[7]
            
            fecha = datetime.datetime.strptime(fecha_str, "%d-%m-%Y")
            detalle_lista = ast.literal_eval(detalle_str)
            
            notas[int(folio)] =[fecha,cliente,rfc,correo,detalle_lista,float(monto),ast.literal_eval(estado)]
except FileNotFoundError:
    print("\nEl archivo CSV no existe. No se han cargado notas previas ")

menu = """
╔═══════════════════════╗
║     MENÚ PRINCIPAL    ║
╠═══════════════════════╢
║ 1. Registrar una nota ║
║ 2. Consultas y        ║
║    reportes           ║
║ 3. Cancelar una nota  ║
║ 4. Recuperar nota     ║
║ 5. Salir              ║
╚═══════════════════════╝
"""
menu_consulta = """
╔════════════════════════════╗
║      MENÚ DE CONSULTA      ║
╟────────────────────────────╢
║ 1. Consultar por período.  ║
║ 2. Consultar por folio.    ║
║ 3. Consulta por cliente    ║
╚════════════════════════════╝
"""
generador_folio=0
patron_fecha = r"^\d{2}-\d{2}-\d{4}$"
fecha_actual =datetime.datetime.now()


while True:
    print(menu)
    opcion=input("\nIngrese una opción:\n")
    if opcion=="1":#ANGEL MORALES VENTURA
        generador_folio=len(notas)
        generador_folio+=1
        folio=generador_folio
        
        while True:
            
            fecha_ingresada_str = input("\nFecha de la nota (dd-mm-aaaa): ").strip()
            
            if not fecha_ingresada_str:
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
            elif not re.match(patron_fecha, fecha_ingresada_str):
                print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
            else:
                try:
                    fecha_ingresada = datetime.datetime.strptime(fecha_ingresada_str, "%d-%m-%Y")
                    if fecha_ingresada > fecha_actual:
                        print("LA FECHA NO DEBE SER POSTERIOR A LA FECHA ACTUAL DEL SISTEMA")
                    else:
                        break
                except ValueError:
                    print("LA FECHA NO ES VÁLIDA/NO EXISTE. INTENTE DENUEVO.")
        
        while True:
            cliente = input("\nNombre del cliente: ").strip()
            
            if cliente== "":
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
            elif any(char.isdigit() for char in cliente):
                print("EL NOMBRE NO PUEDE CONTENER DÍGITOS. INTENTE NUEVAMENTE.")
            else:
                break
        
        while True:
            rfc_ingresado = input("\nIngrese un RFC (por ejemplo: XEXT990101NI4): ").strip().upper()
            
            if not rfc_ingresado:
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
            elif not re.match(r'^[A-Z]{4}[0-9]{6}[A-Z0-9]{3}$', rfc_ingresado):
                print("EL RFC INGRESADO NO TIENE EL FORMATO CORRECTO. INTENTE NUEVAMENTE.")
            else:
                try:
                    fecha_rfc = datetime.datetime.strptime(rfc_ingresado[4:10], '%y%m%d')
                    break
                except ValueError:
                    print("LA FECHA EN EL RFC NO ES VÁLIDA. INTENTE NUEVAMENTE.")
        
        while True:
            correo = input("\nIngrese su correo electrónico gmail (por ejemplo: correo123@gmail.com): ").strip()

            if not correo:
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
            elif not correo.endswith('@gmail.com'):
                print("EL CORREO ELECTRÓNICO DEBE SER DE DOMINIO 'gmail.com'. INTENTE NUEVAMENTE")  
            else:
                break
        
        detalle=list()
        monto=0.0
        notas[folio]=[fecha_ingresada,cliente,rfc_ingresado,correo,detalle,monto,True]
        
        while True:
            
            nombre_servicio = input("\nNombre del servicio: ").strip()
            if nombre_servicio == "":
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                continue

            while True:
                
                costo_servicio = input("\nCosto del servicio: ").strip()
                if costo_servicio == "":
                    print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                    continue
                
                try:
                    costo_servicio = float(costo_servicio)
                    if costo_servicio <= 0:
                        print("EL COSTO DEBE SER MAYOR A 0 PESOS. INTENTE DENUEVO")
                        continue
                    else:
                        notas[folio][4].append((nombre_servicio, costo_servicio))
                        notas[folio][5] += costo_servicio
                        break
                except ValueError:
                    print("SE INGRESÓ UN CARÁCTER NO NUMÉRICO. INTENTE DENUEVO ")
        
            if input("\n¿Agregar otro servicio? (s/n):\n ").lower() != "s":
                print("\n**************************")
                print("         NOTA")
                print("**************************")
                print(f"Folio: {folio:04}")
                print(f"Fecha de la nota: {fecha_ingresada.strftime('%d-%m-%Y')}")
                print(f"Cliente: {cliente}")
                print(f"RFC: {rfc_ingresado}")
                print(f"Correo electrónico: {correo}")
                print("\nDetalles de los servicios:")
                for servicio, costo in notas[folio][4]:
                    print(f"  - {servicio:<20}: ${costo:.2f}")
                print("------------------------------")
                print(f"Monto total:          ${notas[folio][5]:.2f}")
                print("**************************")
                break
    elif opcion=="2":#MONTERO CASTILLO DAVID EDUARDO
        while True:
            print(menu_consulta)
            opcion_consulta=input("\nIngrese una opción/ 0 para volver al menú principal.").strip()
            if opcion_consulta=="0":
                break
            elif opcion_consulta=="1":
                
                while True:
                    print("\nDejar en blanco para usar 01-01-2000")
                    fecha_inicial_str = input("Fecha inicial (dd-mm-aaaa):\n ").strip()
                
                    if fecha_inicial_str =="":
                        print("Se utilizará la fecha por defecto: 01-01-2000.")
                        fecha_inicial = datetime.datetime(2000, 1, 1)
                        break
                    elif not re.match(patron_fecha, fecha_inicial_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, "%d-%m-%Y")
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                    
                while True:
                    print("\nDejar en blanco para usar la fecha actual del sistema.")       
                    fecha_final_str = input("Fecha final (dd-mm-aaaa):\n ")
                    
                    if fecha_final_str =="":
                        print("Se utilizará la fecha actual del sistema.")
                        fecha_final=fecha_actual
                        break
                    elif not re.match(patron_fecha, fecha_final_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    
                    try:
                        fecha_final = datetime.datetime.strptime(fecha_final_str, "%d-%m-%Y")
                        if not fecha_final>=fecha_inicial:
                            print("LA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL.INTENTE NUEVAMENTE.")
                            continue
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                
                # Encabezado de la tabla
                print(f"\n{'Folio':<10}| {'Fecha':<12}| {'Cliente':<20}| {'RFC':<15}| {'Correo':<30}| {'Monto':<10}")
                contador_notas_en_periodo = 0

                for folio, nota in notas.items():
                    if fecha_inicial <= nota[0] <= fecha_final and nota[6]:
                        contador_notas_en_periodo += 1
                        fecha=nota[0]
                        cliente=nota[1]
                        rfc=nota[2]
                        correo=nota[3]
                        monto=nota[5]
                        print(f"{folio:<10}| {fecha.strftime('%d-%m-%Y'):<12}| {cliente:<20}| {rfc:<15}| {correo:<30}| ${monto:>10.2f}")
                        print("-"*90)

                if contador_notas_en_periodo == 0:
                    print("\nNO SE ENCONTRARON NOTAS EN EL PERÍODO ESPECIFICADO.")
                else:
                    print(f"\nTotal de notas encontradas en el período: {contador_notas_en_periodo}")
            elif opcion_consulta=="2":
                    while True:
                        
                        folio_consulta = input("\nIngrese el folio de la nota a consultar: ").strip()
                    
                        if folio_consulta=="":
                            print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                            continue
                        try:
                            folio_consulta=int(folio_consulta)
                        except Exception:
                            print("CARÁCTER NO VALIDO. SOLO DÍGITOS NUMÉRICOS.")
                            continue
                        else:
                            break
                    
                    if folio_consulta in notas and notas[folio_consulta][6]==True:
                        print("\n**************************")
                        print("         NOTA")
                        print("**************************")
                        print(f"Folio: {folio_consulta:04}")
                        print(f"Fecha de la nota: {notas[folio_consulta][0].strftime('%d-%m-%Y')}")
                        print(f"Cliente: {notas[folio_consulta][1]}")
                        print(f"RFC: {notas[folio_consulta][2]}")
                        print(f"Correo electrónico: {notas[folio_consulta][3]}")
                        print("\nDetalles de los servicios:")
                        for servicio, costo in notas[folio_consulta][4]:
                            print(f"  - {servicio:<20}: ${costo:.2f}")
                        print("------------------------------")
                        print(f"Monto total:          ${notas[folio_consulta][5]:.2f}")
                        print("**************************")
                    else:
                        print("NOTA NO ENCONTRADA EN EL SISTEMA.")
                        
            elif opcion_consulta=="3":
                
                rfc_folio_list = []

                for folio, nota in notas.items():
                    if nota[6]:
                        rfc_folio_list.append((nota[2], folio))

                if len(rfc_folio_list) == 0:
                    print("NO HAY RFCs EN EL SISTEMA.")
                else:
                    rfc_folio_list.sort()
                    print(f"\n{'RFC':<15}|{'Folio':<10}")
                    print("-" * 25)
                    for rfc, folio in rfc_folio_list:
                        print(f"{rfc:<15}|{folio:<10}")  
                while True:        
                    consulta_RFC=input("Folio correspondiente al RFC a consultar(0 para volver al menú principal) :\n").strip()
                    
                    if consulta_RFC=="":
                        print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                    elif consulta_RFC=="0":
                        break
                    else:
                        try:
                            consulta_RFC=int(consulta_RFC)
                        except Exception:
                            print("CARÁCTER NO VALIDO. SOLO DÍGITOS NUMÉRICOS.")
                            continue
                            
                        
                    if consulta_RFC in notas and notas[consulta_RFC][6]==True:
                        print("\n**************************")
                        print("         NOTA")
                        print("**************************")
                        print(f"Folio: {consulta_RFC:04}")
                        print(f"Fecha de la nota: {notas[consulta_RFC][0].strftime('%d-%m-%Y')}")
                        print(f"Cliente: {notas[consulta_RFC][1]}")
                        print(f"RFC: {notas[consulta_RFC][2]}")
                        print(f"Correo electrónico: {notas[consulta_RFC][3]}")
                        print("\nDetalles de los servicios:")
                        for servicio, costo in notas[consulta_RFC][4]:
                            print(f"  - {servicio:<20}: ${costo:.2f}")
                        print("------------------------------")
                        print(f"Monto total:          ${notas[consulta_RFC][5]:.2f}")
                        print("**************************\n")
                        while True:
                            pregunta_excel=input("Desea exportar dicha información a un archivo de Excel? (N/S)").upper()
                            if pregunta_excel=="N":
                                print("NO SE EXPORTO DICHA INFORMACIÓN A UN ARCHIVO DE EXCEL")
                                break
                            elif pregunta_excel=="":
                                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO")
                            elif pregunta_excel=="S":
                                from openpyxl.styles import Alignment, Font, Border, Side

                                # Obtener la fecha actual en formato Excel
                                fecha_actual_excel = datetime.datetime.today().date()

                                # Crear un nuevo libro de trabajo
                                libro = openpyxl.Workbook()
                                hoja = libro.active
                                hoja.title = "NOTA"

                                # Establecer el ancho de las columnas
                                hoja.column_dimensions["A"].width = 30
                                hoja.column_dimensions["B"].width = 30
                                
                                fecha_basica=notas[consulta_RFC][0]
                                fecha_basica=fecha_basica.strftime('%d-%m-%Y')

                                # Información básica
                                informacion_basica = {
                                    "Fecha:": fecha_basica,
                                    "Cliente:": notas[consulta_RFC][1],
                                    "RFC:": notas[consulta_RFC][2],
                                    "Correo:": notas[consulta_RFC][3]
                                }

                                # Escribir información básica
                                for i, (etiqueta, valor) in enumerate(informacion_basica.items(), start=1):
                                    celda_etiqueta = hoja[f"A{i}"]
                                    celda_etiqueta.value = etiqueta
                                    celda_etiqueta.font = Font(bold=True)
                                    celda_etiqueta.alignment = Alignment(horizontal='center', vertical='center')

                                    celda_valor = hoja[f"B{i}"]
                                    celda_valor.value = valor
                                    celda_valor.alignment = Alignment(horizontal='center', vertical='center')

                                # Encabezados para los servicios
                                celda_servicio = hoja["A6"]
                                celda_servicio.value = "Servicio"
                                celda_servicio.font = Font(bold=True)
                                celda_servicio.alignment = Alignment(horizontal='center', vertical='center')

                                celda_costo = hoja["B6"]
                                celda_costo.value = "Costo"
                                celda_costo.font = Font(bold=True)
                                celda_costo.alignment = Alignment(horizontal='center', vertical='center')

                                # Escribir detalles de los servicios
                                fila_servicios = 7
                                for servicio, costo in notas[consulta_RFC][4]:
                                    hoja.cell(row=fila_servicios, column=1).value = servicio
                                    hoja.cell(row=fila_servicios, column=2).value = costo
                                    fila_servicios += 1

                                # Escribir monto total
                                celda_monto_total = hoja.cell(row=fila_servicios, column=1)
                                celda_monto_total.value = "Monto total"
                                celda_monto_total.font = Font(bold=True)
                                celda_monto_total.alignment = Alignment(horizontal='center', vertical='center')

                                celda_monto_total_valor = hoja.cell(row=fila_servicios, column=2)
                                celda_monto_total_valor.value = notas[consulta_RFC][5]
                                celda_monto_total_valor.alignment = Alignment(horizontal='center', vertical='center')


                                # Establecer bordes
                                for row in hoja.iter_rows():
                                    for cell in row:
                                        cell.border = Border(bottom=Side(style='thin'))

                                # Nombre del archivo
                                nombre_archivo_excel = f"{notas[consulta_RFC][2]}_{fecha_actual_excel.strftime('%d-%m-%Y')}.xlsx"

                                # Guardar el archivo
                                libro.save(nombre_archivo_excel)
                                print(f"Archivo guardado como {nombre_archivo_excel}.")
                                break
                            else:
                                "OPCIÓN NO VALIDA.INTENTE NUEVAMENTE."     
                    else:
                        print("NOTA NO ENCONTRADA EN EL SISTEMA.")
                        break
                    break
            else:
                print("OPCIÓN NO VALIDA.INTENTE NUEVAMENTE.")

    elif opcion=="3":#GONZALEZ INFANTE ALAN JAIR
        while True:
            folio_cancelar = input("\nIngrese el folio de la nota a cancelar/ 0 para ingresar al menú principal: ").strip()
            
            if folio_cancelar=="":
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                continue

            if folio_cancelar=="0":
                break

            try:
                folio_cancelar=int(folio_cancelar)
                
                if not folio_cancelar in notas:
                    print("\nNOTA NO ENCONTRADA EN EL SISTEMA.")
                    continue
            except Exception:
                print("CARÁCTER NO VÁLIDO. SOLO DÍGITOS NUMÉRICOS")
                continue

            if notas [folio_cancelar][6]==True:
                print("\n**************************")
                print("         NOTA")
                print("**************************")
                print(f"Folio: {folio_cancelar:04}")
                print(f"Fecha de la nota: {notas[folio_cancelar][0].strftime('%d-%m-%Y')}")
                print(f"Cliente: {notas[folio_cancelar][1]}")
                print(f"RFC: {notas[folio_cancelar][2]}")
                print(f"Correo electrónico: {notas[folio_cancelar][3]}")
                print("\nDetalles de los servicios:")
                for servicio, costo in notas[folio_cancelar][4]:
                    print(f"  - {servicio:<20}: ${costo:.2f}")
                print("-------------------------------")
                print(f"Monto total:      ${notas[folio_cancelar][5]:.2f}")
                print("**********************")
            else:
                print("NOTA NO ENCONTRADA EN EL SISTEMA.")
                continue
            while True:

                confirmacion = input("¿Confirmar la cancelación de esta nota? (s/n): ").lower().strip()
                
                if confirmacion=="":
                    
                    print("EL DATO NO PUEDE OMITIRSE. INTENTE NUEVAMENTE.")
                elif confirmacion== "s":

                    notas[folio_cancelar][6] = False
                    print(f"\nNota con folio {folio_cancelar} ha sido cancelada.\n")
                    break
                elif confirmacion=="n":
                    
                    print("\nNOTA NO CANCELADA\n")
                    break
                else:
                    
                    print("OPCIÓN NO VÁLIDA. INTENTE NUEVAMENTE.")
            break
        
    elif opcion=="4":#GONZALEZ INFANTE ALAN JAIR
        encontradas_canceladas = False

        for folio, nota in notas.items():
            if nota[6] == False:
                encontradas_canceladas = True

        if not encontradas_canceladas:
            print("\nNO SE ENCONTRARON NOTAS EN EL SISTEMA.\n")
            continue
        print("Folio   |   Fecha        |   Cliente              |   RFC          |   Correo                        |   Monto")
        print("-" * 100)

        for folio, nota in notas.items():
            if nota[6]==False:
                fecha = nota[0].strftime('%d-%m-%Y')
                cliente = nota[1]
                rfc = nota[2]
                correo = nota[5]
                monto = nota[5]
                
                cliente = clienteljust(23)
                rfc = rfc.ljust(15)
                correo = correo.ljust(30)
                
                print(f"{folio:<7} |   {fecha:<12} |   {cliente} |   {rfc} |   {correo} |   ${monto:.2f}")

        while True:

            folio_rescate = input("\nNota a recuperar(Indique número de folio)/0 para volver al menú principal:\n ").strip()
            
            if folio_rescate=="":
                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                continue
                
            elif folio_rescate=="0":
                break
            else:
                try:
                    folio_rescate=int(folio_rescate)
                    if not folio_rescate in notas:
                        print("NOTA NO ENCONTRADA. INTENTE NUEVAMENTE")
                        continue
                except ValueError:
                    print("Ingrese un número de folio válido.\n")
                    continue
            if folio_rescate in notas and notas[folio_rescate][6]==False:
                print("\n**************************")
                print("         NOTA")
                print("**************************")
                print(f"Folio: {folio_rescate:04}")
                print(f"Fecha de la nota: {notas[folio_rescate][0].strftime('%d-%m-%Y')}")
                print(f"Cliente: {notas[folio_rescate][1]}")
                print(f"RFC: {notas[folio_rescate][2]}")
                print(f"Correo electrónico: {notas[folio_rescate][3]}")
                print("\nDetalles de los servicios.:")
                for servicio, costo in notas[folio_rescate][4]:
                    print(f"  - {servicio:<20}: ${costo:.2f}")
                print("**************************")
                while True:
                    confirmacion = input("¿Confirmar la recuperación de esta nota? (s/n): ").lower().strip()
                    
                    if confirmacion=="":
                     print("NO SE PUEDE OMITIR EL DATO")
                     continue
                    elif confirmacion=="n":
                        print("\nNOTA NO FUE RECUPERADA\n")
                        break
                    elif confirmacion=="s":
                        notas[folio_rescate][6] = True
                        print(f"\nNota con folio {folio_rescate:04} ha sido recuperada.\n")
                        break
                    else:
                        print("OPCIÓN NO VÁLIDA")
                        continue
                break
            else:
                print("FOLIO NO ENCONTRADO EN EL SISTEMA")
        
    elif opcion=="5": #ANGEL MORALES VENTURA
        if input("SEGURO DESEA SALIR DEL PROGRAMA? (S/N o Enter para volver a menu principal):\n ").upper()=="S":
            break
        else:
            continue 
    else:
        print("OPCIÓN NO VALIDA.INTENTE NUEVAMENTE.")
        
nombre_archivo = "notas.csv"
try:
    with open(nombre_archivo,"w",newline='') as archivo_csv:
        escritor = csv.writer(archivo_csv)
        for folio, datos in notas.items():
            fecha,cliente,rfc,correo,detalle,monto,estado = datos
            fecha_str = fecha.strftime('%d-%m-%Y')
            estado_str = estado
            escritor.writerow([folio,fecha_str,cliente,rfc,correo,detalle,monto,estado_str])
    print(f"Se han guardado los datos en {nombre_archivo}")
except Exception as e:
    print(f"ERROR AL GUARDAR LOS DATOS EN EL ARCHIVO CSV: {e}")
        
        
            
        
